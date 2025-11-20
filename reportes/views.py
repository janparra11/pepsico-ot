from datetime import timedelta
from django.utils import timezone
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum, Avg, F, ExpressionWrapper, DurationField, Max

from core.auth import require_roles
from core.roles import Rol
from core.models import AuditLog, Config
from ot.models import OrdenTrabajo, EstadoOT
from inventario.models import MovimientoStock

from openpyxl.formatting.rule import DataBarRule
from django.conf import settings
import os



# ---------- estáticos ----------
def _static_logo_path():
    """
    Devuelve la ruta absoluta del logo si existe.
    Busca en STATIC_ROOT (producción) o en STATICFILES_DIRS (desarrollo).
    """
    filename = os.path.join("img", "logo_pepsico.png")

    # 1) Si tienes STATIC_ROOT (cuando corres collectstatic)
    static_root = getattr(settings, "STATIC_ROOT", None)
    if static_root:
        path = os.path.join(static_root, filename)
        if os.path.exists(path):
            return path

    # 2) Si usas STATICFILES_DIRS en desarrollo
    for base in getattr(settings, "STATICFILES_DIRS", []):
        path = os.path.join(base, filename)
        if os.path.exists(path):
            return path

    # 3) Fallback: BASE_DIR/static/img/logo_pepsico.png
    base_dir_static = os.path.join(settings.BASE_DIR, "static", filename)
    if os.path.exists(base_dir_static):
        return base_dir_static

    # Si no se encuentra, devolvemos None (o podrías lanzar una excepción)
    return None


# ---------- utilidades ----------
def _nombre_archivo_reporte(request, base, filtros=None):
    hoy = timezone.localdate().strftime("%Y%m%d")
    estado = (request.GET.get("estado") or "todos").lower()
    extra = ""
    if filtros:
        fi = (filtros.get("fini") or "").replace("-", "")
        ff = (filtros.get("ffin") or "").replace("-", "")
        if fi or ff:
            extra += f"_{fi or 'all'}-{ff or 'all'}"
        rango = filtros.get("rango") or ""
        if rango:
            extra += f"_{rango}"
    return f"{base}_{hoy}_{estado}{extra}"


def _to_naive(dt):
    if not dt:
        return ""
    return timezone.localtime(dt).replace(tzinfo=None)


def _stats_dashboard(qs_ot, global_top_veh=False, veh_metric="ots"):
    """
    Calcula KPIs y tops.
    - Top repuestos: clave 'total' (Sum(cantidad) o Count(id)).
    - Top vehículos: según 'veh_metric' ('ots' | 'cerradas') y si respeta filtros.
    """
    # KPI
    abiertas = qs_ot.filter(activa=True).count()
    cerradas = qs_ot.filter(activa=False).count()

    horas = []
    for ot in qs_ot.filter(activa=False, fecha_cierre__isnull=False).only("fecha_ingreso", "fecha_cierre"):
        diff = (ot.fecha_cierre - ot.fecha_ingreso).total_seconds() / 3600.0
        if diff >= 0:
            horas.append(diff)
    t_promedio = round(sum(horas) / len(horas), 2) if horas else 0.0

    # Top repuestos (SIEMPRE por salidas)
    fields = [f.name for f in MovimientoStock._meta.get_fields()]
    agg = Sum("cantidad", default=0) if "cantidad" in fields else Count("id")
    top_repuestos = (
        MovimientoStock.objects
        .filter(tipo=MovimientoStock.SALIDA)
        .values("repuesto__codigo", "repuesto__descripcion")
        .annotate(total=agg)
        .order_by("-total")
    )

    # Top vehículos
    base_veh = OrdenTrabajo.objects.all() if global_top_veh else qs_ot
    if veh_metric == "cerradas":
        base_veh = base_veh.filter(activa=False)
    top_vehiculos = (
        base_veh.values("vehiculo__patente")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    return {
        "kpi_abiertas": abiertas,
        "kpi_cerradas": cerradas,
        "kpi_tiempo_prom": t_promedio,
        "top_repuestos": top_repuestos,
        "top_vehiculos": top_vehiculos,
    }


# --- Filtros utilitarios ---
def _parse_filters(request):
    f_ini = (request.GET.get("fini") or "").strip()
    f_fin = (request.GET.get("ffin") or "").strip()
    estado = (request.GET.get("estado") or "").strip()
    taller = (request.GET.get("taller") or "").strip()
    mecanico_id = (request.GET.get("mecanico") or "").strip()
    rango = (request.GET.get("rango") or "").strip()
    veh_metric = (request.GET.get("veh_metric") or "ots").strip()

    # Rango rápido si no vienen fechas
    today = timezone.localdate()
    if rango and not (f_ini or f_fin):
        if rango == "hoy":
            f_ini = f_fin = today.strftime("%Y-%m-%d")
        elif rango == "ult7":
            f_ini = (today - timedelta(days=6)).strftime("%Y-%m-%d")
            f_fin = today.strftime("%Y-%m-%d")
        elif rango == "mes":
            f_ini = today.replace(day=1).strftime("%Y-%m-%d")
            f_fin = today.strftime("%Y-%m-%d")

    qs_ot = OrdenTrabajo.objects.all()
    if f_ini:
        qs_ot = qs_ot.filter(fecha_ingreso__date__gte=f_ini)
    if f_fin:
        qs_ot = qs_ot.filter(fecha_ingreso__date__lte=f_fin)
    if estado:
        qs_ot = qs_ot.filter(estado_actual=estado)
    if taller:
        qs_ot = qs_ot.filter(taller_id=taller)
    if mecanico_id:
        qs_ot = qs_ot.filter(mecanico_asignado_id=mecanico_id)

    return qs_ot, {
        "fini": f_ini, "ffin": f_fin, "estado": estado, "taller": taller,
        "mecanico": mecanico_id, "rango": rango, "veh_metric": veh_metric
    }


# ---------- vistas ----------
@login_required
@require_roles(Rol.SUPERVISOR, Rol.JEFE_TALLER, Rol.ADMIN)
def dashboard_reportes(request):
    qs_ot, filtros = _parse_filters(request)
    stats = _stats_dashboard(qs_ot, global_top_veh=False, veh_metric=filtros.get("veh_metric") or "ots")
    objetivo_h = _sla_objetivo_horas()
    sla_total, sla_dentro, sla_pct = _sla_resumen(qs_ot, objetivo_h)
    sla_mecanicos = _sla_por_mecanico(qs_ot, objetivo_h, limit=10)
    mttr_h = _mttr_horas(qs_ot)
    mttr_mecanicos = _mttr_por_mecanico(qs_ot, limit=10)
    cfg = Config.get_solo()

    # Paginaciones
    page_r = request.GET.get("page_r") or 1
    page_v = request.GET.get("page_v") or 1
    page_o = request.GET.get("page_o") or 1

    rep_paginator = Paginator(stats["top_repuestos"], 20)
    veh_paginator = Paginator(stats["top_vehiculos"], 20)
    top_repuestos_page = rep_paginator.get_page(page_r)
    top_vehiculos_page = veh_paginator.get_page(page_v)

    ots_qs = qs_ot.select_related("vehiculo", "taller").only(
        "folio", "estado_actual", "activa", "fecha_ingreso", "fecha_cierre",
        "vehiculo__patente", "taller__nombre"
    ).order_by("-fecha_ingreso")
    ots_paginator = Paginator(ots_qs, 15)
    ots_page = ots_paginator.get_page(page_o)

    # KPI vencidas (SLA)
    ahora = timezone.now()
    limite = ahora - timedelta(hours=getattr(cfg, "sla_horas", 0) or 0)
    kpi_vencidas = qs_ot.filter(activa=True, fecha_ingreso__lte=limite).count()

    # selects
    from taller.models import Taller
    from django.contrib.auth.models import User
    talleres = Taller.objects.all().order_by("nombre")
    mecanicos = User.objects.filter(is_active=True).order_by("username")

    # IDs vencidas para marcar en tabla
    vencidas_ids = set(
        qs_ot.filter(activa=True, fecha_ingreso__lte=limite).values_list("id", flat=True)
    )

    ctx = {
        "cfg": cfg,
        "filtros": filtros,
        "estados": EstadoOT.choices,
        "talleres": talleres,
        "mecanicos": mecanicos,

        "kpi_abiertas": stats["kpi_abiertas"],
        "kpi_cerradas": stats["kpi_cerradas"],
        "kpi_tiempo_prom": stats["kpi_tiempo_prom"],
        "kpi_vencidas": kpi_vencidas,

        "top_repuestos": top_repuestos_page,
        "top_vehiculos": top_vehiculos_page,
        "ots_page": ots_page,

        "kpi_mttr": mttr_h,
        "mttr_mecanicos": mttr_mecanicos,
        "sla_objetivo": objetivo_h,
        "sla_total": sla_total,
        "sla_dentro": sla_dentro,
        "sla_pct": sla_pct,
        "sla_mecanicos": sla_mecanicos,

        "ots_vencidas_ids": vencidas_ids,
    }
    return render(request, "reportes_dashboard.html", ctx)


# --- Exportaciones ---
@login_required
@require_roles(Rol.SUPERVISOR, Rol.JEFE_TALLER, Rol.ADMIN)
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    qs_ot, filtros = _parse_filters(request)
    stats = _stats_dashboard(qs_ot, veh_metric=filtros.get("veh_metric") or "ots")
    cfg = Config.get_solo()

    wb = Workbook()

    # Hoja 1: Resumen
    ws = wb.active
    ws.title = "Resumen"

    for col, w in zip("ABCDEFGH", [16, 20, 26, 16, 18, 18, 14, 14]):
        ws.column_dimensions[col].width = w

    # Encabezado (logo si existe) — SIN bloques duplicados
    logo_path = _static_logo_path()
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 140
            img.height = 40
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 32
            ws.row_dimensions[2].height = 22
            ws.merge_cells("C1:G1"); ws["C1"] = f"{cfg.nombre_taller} - Reporte de Órdenes de Trabajo"
            ws.merge_cells("C2:G2"); ws["C2"] = f"Generado: {timezone.localtime(timezone.now()).strftime('%d-%m-%Y %H:%M')}"
        except Exception:
            ws.append([f"{cfg.nombre_taller} - Reporte de Órdenes de Trabajo"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            ws.append([f"Generado: {timezone.localtime(timezone.now()).strftime('%d-%m-%Y %H:%M')}"])
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    else:
        ws.append([f"{cfg.nombre_taller} - Reporte de Órdenes de Trabajo"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws.append([f"Generado: {timezone.localtime(timezone.now()).strftime('%d-%m-%Y %H:%M')}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    ws.append([])  # línea en blanco

    # Filtros
    ws.append(["Filtros"])
    if filtros.get("rango"):
        ws.append(["Rango rápido", {"hoy": "Hoy", "ult7": "Últimos 7 días", "mes": "Este mes"}.get(filtros["rango"], filtros["rango"])])
    ws.append(["Desde", filtros["fini"] or "(todos)"])
    ws.append(["Hasta", filtros["ffin"] or "(todos)"])
    ws.append(["Estado", filtros["estado"] or "(todos)"])
    ws.append(["Taller", filtros["taller"] or "(todos)"])
    ws.append(["Mecánico", filtros["mecanico"] or "(todos)"])
    ws.append([])

    # KPIs
    ws.append(["KPIs"])
    ws.append(["OTs abiertas", stats["kpi_abiertas"]])
    ws.append(["OTs cerradas", stats["kpi_cerradas"]])
    ws.append(["Tiempo prom. reparación (h)", stats["kpi_tiempo_prom"]])

    # SLA y vencidas
    ahora = timezone.now()
    limite = ahora - timedelta(hours=getattr(cfg, "sla_horas", 0) or 0)
    vencidas = qs_ot.filter(activa=True, fecha_ingreso__lte=limite).count()
    ws.append(["SLA (horas)", getattr(cfg, "sla_horas", 0) or 0])
    ws.append([f"OTs vencidas (>{getattr(cfg, 'sla_horas', 0) or 0}h abiertas)", vencidas])

    # Hoja 2: OTs
    ws2 = wb.create_sheet("OTs")
    headers = ["Folio", "Patente", "Estado", "Taller", "Ingreso", "Cierre", "Activa"]
    ws2.append(headers)

    for ot in qs_ot.select_related("vehiculo", "taller").only(
        "folio", "estado_actual", "activa", "fecha_ingreso", "fecha_cierre",
        "vehiculo__patente", "taller__nombre"
    ):
        ws2.append([
            ot.folio,
            getattr(ot.vehiculo, "patente", ""),
            ot.get_estado_actual_display(),
            getattr(ot.taller, "nombre", ""),
            _to_naive(ot.fecha_ingreso),
            _to_naive(ot.fecha_cierre),
            "Sí" if ot.activa else "No"
        ])

    for i in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 18
    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            if cell.value:
                cell.number_format = "yyyy-mm-dd hh:mm"

    # Hoja 3: Top Repuestos
    ws3 = wb.create_sheet("Top Repuestos")
    ws3.append(["Código", "Descripción", "Total"])
    for r in stats["top_repuestos"][:50]:
        ws3.append([r["repuesto__codigo"], r["repuesto__descripcion"], r["total"]])
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 14
    for cell in ws3["C"][1:]:
        cell.number_format = "#,##0.##"
    if ws3.max_row > 1:
        ws3.conditional_formatting.add(
            f"C2:C{ws3.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="638EC6")
        )

    # Hoja 4: Top Vehículos
    hoja_veh = "Top Vehículos (OTs)" if (filtros.get("veh_metric") or "ots") == "ots" else "Top Vehículos (Cerradas)"
    ws4 = wb.create_sheet(hoja_veh)
    ws4.append(["Patente", "OTs"])
    for v in stats["top_vehiculos"][:50]:
        ws4.append([v["vehiculo__patente"], v["total"]])
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 10
    for cell in ws4["B"][1:]:
        cell.number_format = "#,##0"
    if ws4.max_row > 1:
        ws4.conditional_formatting.add(
            f"B2:B{ws4.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="95C76F")
        )

    # Hoja 5: Resumen por estado
    ws5 = wb.create_sheet("Resumen por estado")
    ws5.append(["Estado", "OTs", "Tiempo prom. (h)"])
    cerradas = qs_ot.filter(
        activa=False,
        fecha_cierre__isnull=False,
        fecha_cierre__gte=F("fecha_ingreso")
    ).annotate(
        dur_td=ExpressionWrapper(F("fecha_cierre") - F("fecha_ingreso"), output_field=DurationField())
    )
    resumen = (
        cerradas.values("estado_actual")
        .annotate(total=Count("id"), prom_td=Avg("dur_td"))
        .order_by("estado_actual")
    )
    estado_map = dict(EstadoOT.choices)
    for r in resumen:
        prom_h = round(r["prom_td"].total_seconds() / 3600.0, 2) if r["prom_td"] else ""
        ws5.append([estado_map.get(r["estado_actual"], r["estado_actual"]), r["total"], prom_h])
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 10
    ws5.column_dimensions["C"].width = 18
    for cell in ws5["B"][2:]:
        cell.number_format = "#,##0"
    for cell in ws5["C"][2:]:
        cell.number_format = "#,##0.00"

    # Hoja 6: MTTR por mecánico
    wsM = wb.create_sheet("MTTR por mecánico")
    wsM.append(["MTTR general (h):", _mttr_horas(qs_ot)])
    wsM.append([])
    wsM.append(["Mecánico", "MTTR (h)"])
    for user, h in _mttr_por_mecanico(qs_ot, limit=100):
        wsM.append([user, h if h is not None else ""])
    wsM.column_dimensions["A"].width = 28
    wsM.column_dimensions["B"].width = 14
    for cell in wsM["B"][3:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"

    # Hoja 7: Resumen por taller
    cerradas_ok = qs_ot.filter(
        activa=False,
        fecha_cierre__isnull=False,
        fecha_cierre__gte=F("fecha_ingreso"),
    ).annotate(
        dur_td=ExpressionWrapper(F("fecha_cierre") - F("fecha_ingreso"), output_field=DurationField())
    )
    agg_base = qs_ot.values("taller__nombre").annotate(
        total_ots=Count("id"),
        abiertas=Count("id", filter=Q(activa=True)),
        cerradas=Count("id", filter=Q(activa=False)),
    ).order_by("taller__nombre")
    dur_prom_map = {
        r["taller__nombre"] or "—": (r["avg_dur"].total_seconds() / 3600.0) if r["avg_dur"] else 0.0
        for r in cerradas_ok.values("taller__nombre").annotate(avg_dur=Avg("dur_td"))
    }
    wsT = wb.create_sheet("Resumen por taller")
    wsT.append(["Taller", "OTs totales", "Abiertas", "Cerradas", "Tiempo prom. (h)"])
    for row in agg_base:
        nombre = row["taller__nombre"] or "—"
        wsT.append([nombre, row["total_ots"], row["abiertas"], row["cerradas"], round(dur_prom_map.get(nombre, 0.0), 2)])
    from openpyxl.utils import get_column_letter as _gcl
    for i, w in enumerate([30, 14, 12, 12, 18], start=1):
        wsT.column_dimensions[_gcl(i)].width = w
    for rng in ("B", "C", "D"):
        for cell in wsT[rng][1:]:
            cell.number_format = "#,##0"
    for cell in wsT["E"][1:]:
        cell.number_format = "0.00"

    # Hoja 8: SLA
    wsS = wb.create_sheet("SLA")
    obj_h = _sla_objetivo_horas()
    tot, din, pct = _sla_resumen(qs_ot, obj_h)
    wsS.append(["Objetivo SLA (h)", obj_h])
    wsS.append(["OTs cerradas válidas", tot])
    wsS.append(["Cumplen SLA", din])
    wsS.append(["Cumplimiento (%)", pct])
    wsS.append([])
    wsS.append(["Mecánico", "OTs cerradas", "Dentro SLA", "Cumplimiento (%)"])
    for user, t, d, p in _sla_por_mecanico(qs_ot, obj_h, limit=100):
        wsS.append([user, t, d, p])
    wsS.column_dimensions["A"].width = 28
    for col in ("B", "C", "D"):
        wsS.column_dimensions[col].width = 18
    for cell in wsS["D"][7:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.0"

    # nombre archivo y respuesta
    fname = _nombre_archivo_reporte(request, "reporte_ots") + ".xlsx"
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(resp)
    AuditLog.objects.create(app="REPORTES", action="EXPORT_EXCEL", user=request.user, extra=str(request.GET.dict()))
    return resp


@login_required
@require_roles(Rol.SUPERVISOR, Rol.JEFE_TALLER, Rol.ADMIN)
def export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader  # logo

    qs_ot, filtros = _parse_filters(request)
    stats = _stats_dashboard(qs_ot, global_top_veh=False,
                             veh_metric=filtros.get("veh_metric") or "ots")
    cfg = Config.get_solo()

    # === NUEVO: métricas para MTTR/SLA por mecánico ===
    objetivo_h = _sla_objetivo_horas()
    sla_total, sla_dentro, sla_pct = _sla_resumen(qs_ot, objetivo_h)
    sla_mecanicos = _sla_por_mecanico(qs_ot, objetivo_h, limit=10)
    mttr_h = _mttr_horas(qs_ot)
    mttr_mecanicos = _mttr_por_mecanico(qs_ot, limit=10)

    resp = HttpResponse(content_type="application/pdf")
    fname = _nombre_archivo_reporte(request, "reporte_ots", filtros) + ".pdf"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    c = canvas.Canvas(resp, pagesize=A4)
    w, h = A4

    # Cabecera
    y = h - 2 * cm
    left = 2 * cm
    top = y

    logo_path = _static_logo_path()
    logo_w, logo_h = (0, 0)
    if logo_path:
        try:
            img = ImageReader(logo_path)
            max_h = 2.2 * cm
            iw, ih = img.getSize()
            scale = max_h / float(ih)
            logo_w, logo_h = iw * scale, ih * scale
            c.drawImage(
                img,
                left,
                top - logo_h,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            logo_w, logo_h = (0, 0)

    text_x = left + (logo_w + 0.6 * cm if logo_w else 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(
        text_x, top, f"{cfg.nombre_taller} · Reporte de Órdenes de Trabajo"
    )
    c.setFont("Helvetica", 10)
    c.drawString(
        text_x,
        top - 0.6 * cm,
        f"Generado: {timezone.localtime(timezone.now()).strftime('%d-%m-%Y %H:%M')}",
    )
    y = min(top - logo_h, top - 0.6 * cm) - 0.8 * cm

    # Filtros
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left, y, "Filtros:")
    c.setFont("Helvetica", 10)
    y -= 0.5 * cm
    if filtros.get("rango"):
        etiqueta = {
            "hoy": "Hoy",
            "ult7": "Últimos 7 días",
            "mes": "Este mes",
        }.get(filtros["rango"], filtros["rango"])
        c.drawString(left, y, f"Rango rápido: {etiqueta}")
        y -= 0.5 * cm
    c.drawString(
        left,
        y,
        f"Desde: {filtros['fini'] or '(todos)'}   Hasta: {filtros['ffin'] or '(todos)'}",
    )
    y -= 0.5 * cm
    c.drawString(
        left,
        y,
        f"Estado: {filtros['estado'] or '(todos)'}   Taller: {filtros['taller'] or '(todos)'}   Mecánico: {filtros['mecanico'] or '(todos)'}",
    )
    y -= 0.8 * cm

    # KPIs + SLA
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y, "KPIs:")
    c.setFont("Helvetica", 10)
    y -= 0.5 * cm
    c.drawString(
        2 * cm,
        y,
        f"OTs abiertas: {stats['kpi_abiertas']}   "
        f"OTs cerradas: {stats['kpi_cerradas']}   "
        f"Tiempo prom: {stats['kpi_tiempo_prom']} h   "
        f"MTTR: {mttr_h:.2f} h",
    )
    y -= 0.5 * cm

    ahora = timezone.now()
    limite = ahora - timedelta(hours=getattr(cfg, "sla_horas", 0) or 0)
    vencidas = qs_ot.filter(activa=True, fecha_ingreso__lte=limite).count()
    c.drawString(
        2 * cm,
        y,
        f"SLA vencidas (>{getattr(cfg, 'sla_horas', 0) or 0} h abiertas): {vencidas}",
    )
    y -= 0.5 * cm
    c.drawString(
        2 * cm,
        y,
        f"SLA objetivo cerradas (≤ {objetivo_h} h): "
        f"{sla_pct}%  ({sla_dentro}/{sla_total} dentro de SLA)",
    )
    y -= 0.8 * cm

    # Top Repuestos (con “Últ. mov.”)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y, "Top 10 Repuestos (por consumo en filtros)")
    y -= 0.6 * cm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(2 * cm, y, "Código")
    c.drawString(7 * cm, y, "Descripción")
    c.drawRightString(15 * cm, y, "Cantidad")
    c.drawRightString(19 * cm, y, "Últ. mov.")
    y -= 0.45 * cm
    c.setFont("Helvetica", 9)

    ult_mov_map = {
        r["repuesto__codigo"]: r["ultimo"]
        for r in MovimientoStock.objects.filter(tipo=MovimientoStock.SALIDA)
        .values("repuesto__codigo")
        .annotate(ultimo=Max("creado_en"))
    }
    for r in list(stats["top_repuestos"][:10]):
        cod = r["repuesto__codigo"] or "—"
        desc = (r["repuesto__descripcion"] or "—")[:50]
        total = r["total"] or 0
        ult = ult_mov_map.get(cod)
        ult_txt = (
            timezone.localtime(ult).strftime("%d-%m-%Y %H:%M")
            if ult
            else "—"
        )
        c.drawString(2 * cm, y, cod[:20])
        c.drawString(7 * cm, y, desc)
        c.drawRightString(15 * cm, y, f"{total}")
        c.drawRightString(19 * cm, y, ult_txt)
        y -= 0.42 * cm
        if y < 2 * cm:
            c.showPage()
            y = h - 2 * cm
            c.setFont("Helvetica", 9)

    y -= 0.4 * cm

    # Top Vehículos (taller más frecuente + último servicio)
    veh_counts = (
        qs_ot.values("vehiculo__patente")
        .annotate(total=Count("id"))
        .order_by("-total", "vehiculo__patente")[:10]
    )
    patentes = [v["vehiculo__patente"] for v in veh_counts]
    tfreq_qs = (
        qs_ot.filter(vehiculo__patente__in=patentes)
        .values("vehiculo__patente", "taller__nombre")
        .annotate(n=Count("id"))
        .order_by("vehiculo__patente", "-n")
    )
    taller_frecuente = {}
    for row in tfreq_qs:
        p = row["vehiculo__patente"]
        if p not in taller_frecuente:
            taller_frecuente[p] = row["taller__nombre"] or "—"

    ult_serv_qs = (
        qs_ot.filter(vehiculo__patente__in=patentes)
        .values("vehiculo__patente")
        .annotate(ultimo=Max("fecha_ingreso"))
    )
    ultimo_serv = {r["vehiculo__patente"]: r["ultimo"] for r in ult_serv_qs}

    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y, "Top 10 Vehículos (por OTs en filtros)")
    y -= 0.6 * cm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(2 * cm, y, "Patente")
    c.drawRightString(8 * cm, y, "OTs")
    c.drawString(9 * cm, y, "Taller más frecuente")
    c.drawRightString(16 * cm, y, "Últ. servicio")
    c.drawRightString(19 * cm, y, "Hace")
    y -= 0.45 * cm
    c.setFont("Helvetica", 9)

    now_local = timezone.localtime(timezone.now())
    for v in veh_counts:
        p = v["vehiculo__patente"] or "—"
        tot = v["total"] or 0
        tall = taller_frecuente.get(p, "—")
        u = ultimo_serv.get(p)
        if u:
            u_loc = timezone.localtime(u)
            ult_txt = u_loc.strftime("%d-%m-%Y %H:%M")
            dias = (now_local.date() - u_loc.date()).days
            hace_txt = f"{dias} d"
        else:
            ult_txt = "—"
            hace_txt = "—"
        c.drawString(2 * cm, y, p[:15])
        c.drawRightString(8 * cm, y, str(tot))
        c.drawString(9 * cm, y, (tall or "—")[:24])
        c.drawRightString(16 * cm, y, ult_txt)
        c.drawRightString(19 * cm, y, hace_txt)
        y -= 0.42 * cm
        if y < 2 * cm:
            c.showPage()
            y = h - 2 * cm
            c.setFont("Helvetica", 9)

    # ==== NUEVA PÁGINA: MTTR y SLA por mecánico ====
    c.showPage()
    y = h - 2 * cm

    # MTTR por mecánico
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, y, "MTTR por mecánico (Top 10)")
    y -= 0.8 * cm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(2 * cm, y, "Mecánico")
    c.drawRightString(10 * cm, y, "MTTR (h)")
    y -= 0.5 * cm
    c.setFont("Helvetica", 9)

    for user, horas in mttr_mecanicos:
        c.drawString(2 * cm, y, user or "—")
        txt = f"{horas:.2f}" if horas is not None else "—"
        c.drawRightString(10 * cm, y, txt)
        y -= 0.4 * cm
        if y < 2 * cm:
            c.showPage()
            y = h - 2 * cm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(2 * cm, y, "MTTR por mecánico (cont.)")
            y -= 0.8 * cm
            c.setFont("Helvetica-Bold", 9)
            c.drawString(2 * cm, y, "Mecánico")
            c.drawRightString(10 * cm, y, "MTTR (h)")
            y -= 0.5 * cm
            c.setFont("Helvetica", 9)

    # Un poco de espacio antes del SLA
    y -= 0.6 * cm
    if y < 4 * cm:
        c.showPage()
        y = h - 2 * cm

    # SLA por mecánico
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, y, f"SLA por mecánico (≤ {objetivo_h} h)")
    y -= 0.8 * cm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(2 * cm, y, "Mecánico")
    c.drawRightString(11 * cm, y, "Cerradas")
    c.drawRightString(15 * cm, y, "Dentro SLA")
    c.drawRightString(19 * cm, y, "Cumplimiento")
    y -= 0.5 * cm
    c.setFont("Helvetica", 9)

    for user, total, dentro, pct in sla_mecanicos:
        c.drawString(2 * cm, y, user or "—")
        c.drawRightString(11 * cm, y, str(total))
        c.drawRightString(15 * cm, y, str(dentro))
        c.drawRightString(19 * cm, y, f"{pct:.1f}%")
        y -= 0.4 * cm
        if y < 2 * cm:
            c.showPage()
            y = h - 2 * cm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(2 * cm, y, f"SLA por mecánico (≤ {objetivo_h} h) (cont.)")
            y -= 0.8 * cm
            c.setFont("Helvetica-Bold", 9)
            c.drawString(2 * cm, y, "Mecánico")
            c.drawRightString(11 * cm, y, "Cerradas")
            c.drawRightString(15 * cm, y, "Dentro SLA")
            c.drawRightString(19 * cm, y, "Cumplimiento")
            y -= 0.5 * cm
            c.setFont("Helvetica", 9)

    # ==== Resumen por taller (como ya lo tenías) ====
    c.showPage()
    y = h - 2 * cm
    X_TALLER, X_TOTAL, X_ABIERT, X_CERRAD, X_PROM = (
        2 * cm,
        11 * cm,
        13 * cm,
        15 * cm,
        19 * cm,
    )
    c.setFont("Helvetica-Bold", 12)
    c.drawString(X_TALLER, y, "Resumen por taller")
    y -= 0.8 * cm

    cerradas_ok = qs_ot.filter(
        activa=False,
        fecha_cierre__isnull=False,
        fecha_cierre__gte=F("fecha_ingreso"),
    ).annotate(
        dur_td=ExpressionWrapper(
            F("fecha_cierre") - F("fecha_ingreso"),
            output_field=DurationField(),
        )
    )
    agg_base = (
        qs_ot.values("taller__nombre")
        .annotate(
            total_ots=Count("id"),
            abiertas=Count("id", filter=Q(activa=True)),
            cerradas=Count("id", filter=Q(activa=False)),
        )
        .order_by("taller__nombre")
    )
    dur_prom_map = {
        r["taller__nombre"] or "—": (
            r["avg_dur"].total_seconds() / 3600.0 if r["avg_dur"] else 0.0
        )
        for r in cerradas_ok.values("taller__nombre").annotate(
            avg_dur=Avg("dur_td")
        )
    }

    c.setFont("Helvetica-Bold", 9)
    c.drawString(X_TALLER, y, "Taller")
    c.drawRightString(X_TOTAL, y, "Totales")
    c.drawRightString(X_ABIERT, y, "Abiertas")
    c.drawRightString(X_CERRAD, y, "Cerradas")
    c.drawRightString(X_PROM, y, "T. prom (h)")
    y -= 0.5 * cm
    c.setFont("Helvetica", 9)

    def ellipsize(txt, max_len=38):
        if not txt:
            return "—"
        t = str(txt)
        return t if len(t) <= max_len else (t[: max_len - 1] + "…")

    for row in agg_base:
        nombre = ellipsize(row["taller__nombre"] or "—")
        t = row["total_ots"] or 0
        a = row["abiertas"] or 0
        crr = row["cerradas"] or 0
        prom = dur_prom_map.get(row["taller__nombre"] or "—", 0.0)
        c.drawString(X_TALLER, y, nombre)
        c.drawRightString(X_TOTAL, y, f"{t:d}")
        c.drawRightString(X_ABIERT, y, f"{a:d}")
        c.drawRightString(X_CERRAD, y, f"{crr:d}")
        c.drawRightString(X_PROM, y, f"{prom:.2f}")
        y -= 0.45 * cm
        if y < 2 * cm:
            c.showPage()
            y = h - 2 * cm
            c.setFont("Helvetica-Bold", 9)
            c.drawString(X_TALLER, y, "Taller")
            c.drawRightString(X_TOTAL, y, "Totales")
            c.drawRightString(X_ABIERT, y, "Abiertas")
            c.drawRightString(X_CERRAD, y, "Cerradas")
            c.drawRightString(X_PROM, y, "T. prom (h)")
            y -= 0.5 * cm
            c.setFont("Helvetica", 9)

    c.showPage()
    c.save()
    AuditLog.objects.create(
        app="REPORTES",
        action="EXPORT_PDF",
        user=request.user,
        extra=str(request.GET.dict()),
    )
    return resp

# ---------- métricas auxiliares ----------
def _mttr_horas(qs):
    qs_ok = qs.filter(
        activa=False, fecha_cierre__isnull=False, fecha_cierre__gte=F("fecha_ingreso"),
    ).annotate(dur_td=ExpressionWrapper(F("fecha_cierre") - F("fecha_ingreso"), output_field=DurationField()))
    val = qs_ok.aggregate(prom=Avg("dur_td"))["prom"]
    return round(val.total_seconds()/3600.0, 2) if val else 0.0


def _mttr_por_mecanico(qs, limit=10):
    qs_ok = qs.filter(
        activa=False, fecha_cierre__isnull=False, fecha_cierre__gte=F("fecha_ingreso"),
        mecanico_asignado__isnull=False,
    ).annotate(dur_td=ExpressionWrapper(F("fecha_cierre") - F("fecha_ingreso"), output_field=DurationField()))
    data = (qs_ok.values("mecanico_asignado__username").annotate(prom=Avg("dur_td")).order_by("prom")[:limit])
    out = []
    for r in data:
        h = r["prom"].total_seconds()/3600.0 if r["prom"] else None
        out.append((r["mecanico_asignado__username"] or "—", round(h, 2) if h is not None else None))
    return out


def _sla_objetivo_horas():
    try:
        cfg = Config.get_solo()
        return getattr(cfg, "sla_objetivo_horas", 48)
    except Exception:
        return 48


def _cerradas_validas(qs):
    return qs.filter(
        activa=False, fecha_cierre__isnull=False, fecha_cierre__gte=F("fecha_ingreso"),
    ).annotate(dur_td=ExpressionWrapper(F("fecha_cierre") - F("fecha_ingreso"), output_field=DurationField()))


def _sla_resumen(qs, objetivo_h):
    cerr = _cerradas_validas(qs)
    total = cerr.count()
    if not total:
        return (0, 0, 0.0)
    dentro = cerr.filter(dur_td__lte=timedelta(hours=objetivo_h)).count()
    pct = round(dentro * 100.0 / total, 1)
    return (total, dentro, pct)


def _sla_por_mecanico(qs, objetivo_h, limit=20):
    cerr = _cerradas_validas(qs).filter(mecanico_asignado__isnull=False)
    base = (
        cerr.values("mecanico_asignado__username")
            .annotate(total=Count("id"), dentro=Count("id", filter=Q(dur_td__lte=timedelta(hours=objetivo_h))))
            .order_by("-dentro", "-total")
    )
    out = []
    for r in base[:limit]:
        t = r["total"] or 0
        d = r["dentro"] or 0
        pct = round(d * 100.0 / t, 1) if t else 0.0
        out.append((r["mecanico_asignado__username"] or "—", t, d, pct))
    return out
